"""
rvtools_client.py — RVTools auto-run + XLSX parser
====================================================
• Detects if RVTools.exe is installed
• Runs RVTools silently per vCenter (ExportAll2xlsx)
• Scans reports folder + Desktop for existing XLSX exports
• Parses vInfo sheet → rich VM inventory / summary
• Tries to download & install RVTools if missing (requires internet)
"""

import os
import re
import json
import glob
import shutil
import hashlib
import datetime
import subprocess
import urllib.request
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).parent / ".env")

# ── Paths ────────────────────────────────────────────────────────────────────
RVTOOLS_EXE_CANDIDATES = [
    r"C:\Program Files (x86)\RobWare\RVTools\RVTools.exe",
    r"C:\Program Files\RobWare\RVTools\RVTools.exe",
    r"C:\Program Files (x86)\RVTools\RVTools.exe",
    r"C:\Program Files\RVTools\RVTools.exe",
    r"C:\RVTools\RVTools.exe",
]
EXPORTS_DIR = os.getenv("RVTOOLS_EXPORTS_DIR", r"C:\caas-dashboard\rvtools_exports")
INSTALLER_DIR = r"C:\caas-dashboard\rvtools_installer"
RVTOOLS_DOWNLOAD_URL = "https://www.robware.net/rvtools/download/latest"   # best-effort, may fail on air-gapped

SCAN_DIRS = [
    EXPORTS_DIR,
    r"C:\Users\Administrator\Desktop",
    r"C:\Users\Public\Desktop",
]

# Prefer rvtools_exports over Desktop when same file exists in both
SCAN_PRIORITY_DIR = EXPORTS_DIR

# ── vCenter registry (mirrors vmware_client logic) ──────────────────────────
def _vcenter_list():
    raw = os.getenv("VCENTER_HOSTS") or os.getenv("VCENTER_HOST", "")
    hosts = [h.strip() for h in raw.split(",") if h.strip()]
    out = []
    for i, host in enumerate(hosts, 1):
        user = os.getenv(f"VCENTER_USER_{i}") or os.getenv("VCENTER_USER", "administrator@vsphere.local")
        pwd  = os.getenv(f"VCENTER_PASSWORD_{i}") or os.getenv("VCENTER_PASSWORD", "")
        port = int(os.getenv(f"VCENTER_PORT_{i}") or os.getenv("VCENTER_PORT", 443))
        name = os.getenv(f"VCENTER_NAME_{i}") or f"vCenter-{i}"
        out.append({"host": host, "user": user, "pwd": pwd, "port": port, "name": name})
    return out

VCENTERS = _vcenter_list()


# ── RVTools detection ─────────────────────────────────────────────────────────
def find_rvtools_exe():
    for path in RVTOOLS_EXE_CANDIDATES:
        if os.path.isfile(path):
            return path
    return None


def is_rvtools_installed():
    return find_rvtools_exe() is not None


# ── RVTools installer (best-effort, fails gracefully on air-gapped) ──────────
def try_install_rvtools():
    """
    Attempt to download + silently install RVTools.
    Returns {"success": bool, "message": str}
    """
    exe = find_rvtools_exe()
    if exe:
        return {"success": True, "message": f"RVTools already installed at {exe}"}

    os.makedirs(INSTALLER_DIR, exist_ok=True)
    installer_path = os.path.join(INSTALLER_DIR, "RVTools_setup.exe")

    # Try to download
    try:
        print(f"[RVTools] Downloading installer from {RVTOOLS_DOWNLOAD_URL} …")
        req = urllib.request.Request(
            RVTOOLS_DOWNLOAD_URL,
            headers={"User-Agent": "Mozilla/5.0"}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            with open(installer_path, "wb") as fh:
                shutil.copyfileobj(resp, fh)
    except Exception as e:
        return {
            "success": False,
            "message": (
                f"RVTools is not installed and could not be downloaded automatically "
                f"({e}). Please manually install RVTools from https://www.robware.net "
                f"and re-run. Alternatively, place exported .xlsx files in "
                f"{EXPORTS_DIR} and use 'Scan Reports'."
            )
        }

    # Run installer silently
    try:
        result = subprocess.run(
            [installer_path, "/S"],
            capture_output=True, timeout=120
        )
        exe_after = find_rvtools_exe()
        if exe_after:
            return {"success": True, "message": f"RVTools installed successfully at {exe_after}"}
        else:
            return {"success": False, "message": "Installer ran but RVTools.exe not found — installation may need a reboot or manual steps."}
    except Exception as e:
        return {"success": False, "message": f"Installer error: {e}"}


# ── Run RVTools per vCenter ───────────────────────────────────────────────────
def run_rvtools_for_vcenter(vc: dict, output_dir: str = EXPORTS_DIR) -> dict:
    """
    Run RVTools.exe silently for one vCenter → save XLSX in output_dir.
    Returns {"success": bool, "message": str, "file": path or None}
    """
    os.makedirs(output_dir, exist_ok=True)
    exe = find_rvtools_exe()
    if not exe:
        install_result = try_install_rvtools()
        if not install_result["success"]:
            return install_result
        exe = find_rvtools_exe()
        if not exe:
            return {"success": False, "message": "RVTools not found after install attempt.", "file": None}

    safe_name = re.sub(r"[^\w\-]", "_", vc["name"])
    filename  = f"RVTools_{safe_name}_{vc['host'].replace('.','_')}.xlsx"
    out_path  = os.path.join(output_dir, filename)

    cmd = [
        exe,
        "-s", vc["host"],
        "-u", vc["user"],
        "-p", vc["pwd"],
        "-c", "ExportAll2xlsx",
        "-d", output_dir,
        "-f", filename,
    ]
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=300,     # 5 minutes max per vCenter
            text=True
        )
        if os.path.isfile(out_path):
            return {"success": True, "message": f"Export complete: {out_path}", "file": out_path}
        else:
            return {
                "success": False,
                "message": f"RVTools ran but output file not found. stdout={result.stdout[:300]} stderr={result.stderr[:300]}",
                "file": None
            }
    except subprocess.TimeoutExpired:
        return {"success": False, "message": "RVTools timed out (5 min limit).", "file": None}
    except Exception as e:
        return {"success": False, "message": str(e), "file": None}


def run_rvtools_all(output_dir: str = EXPORTS_DIR) -> list:
    """Run RVTools for every configured vCenter. Returns list of results."""
    return [
        {**run_rvtools_for_vcenter(vc, output_dir), "vcenter_host": vc["host"], "vcenter_name": vc["name"]}
        for vc in VCENTERS
    ]


# ── XLSX parser ───────────────────────────────────────────────────────────────
_VINFO_COLS = {
    "VM":           "VM",
    "Powerstate":   "Powerstate",
    "Template":     "Template",
    "CPUs":         "CPUs",
    "Memory":       "Memory",
    "NICs":         "NICs",
    "Disks":        "Disks",
    "Provisioned MB":    "Provisioned MB",
    "In Use MB":         "In Use MB",
    "Unshared MB":       "Unshared MB",
    "Primary IP Address":"Primary IP Address",
    "DNS Name":          "DNS Name",
    "Resource pool":     "Resource pool",
    "Folder":            "Folder",
    "Cluster":           "Cluster",
    "Host":              "Host",
    "OS according to the VMware Tools":  "OS (Tools)",
    "OS according to the configuration file": "OS (Config)",
    "Datacenter":        "Datacenter",
    "VI SDK Server":     "VI SDK Server",
    "VI SDK UUID":       "VI SDK UUID",
    "Annotation":        "Annotation",
    "PowerOn":           "PowerOn",
    "Creation date":     "Creation date",
}


def _parse_vinfo_sheet(ws) -> list:
    """Parse vInfo worksheet → list of VM dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    def g(row, col_name):
        i = col_idx.get(col_name)
        if i is None:
            return None
        v = row[i] if i < len(row) else None
        return v

    vms = []
    for row in rows[1:]:
        name = g(row, "VM")
        if not name:
            continue
        # Build compact VM record
        vm = {
            "name":            str(name),
            "powerstate":      str(g(row, "Powerstate") or "").strip(),
            "template":        str(g(row, "Template") or "").strip().upper() == "TRUE",
            "cpus":            _int(g(row, "CPUs")),
            "memory_mb":       _int(g(row, "Memory")),
            "nics":            _int(g(row, "NICs")),
            "disks":           _int(g(row, "Disks")),
            "provisioned_mb":  _int(g(row, "Provisioned MiB") or g(row, "Provisioned MB")),
            "inuse_mb":        _int(g(row, "In Use MiB") or g(row, "In Use MB")),
            "unshared_mb":     _int(g(row, "Unshared MiB") or g(row, "Unshared MB")),
            "disk_capacity_mb": _int(g(row, "Total disk capacity MiB") or g(row, "Total disk capacity MB")),
            "primary_ip":      str(g(row, "Primary IP Address") or "").strip(),
            "dns_name":        str(g(row, "DNS Name") or "").strip(),
            "resource_pool":   str(g(row, "Resource pool") or "").strip(),
            "folder":          str(g(row, "Folder") or "").strip(),
            "cluster":         str(g(row, "Cluster") or "").strip(),
            "host":            str(g(row, "Host") or "").strip(),
            "os_tools":        str(g(row, "OS according to the VMware Tools") or "").strip(),
            "os_config":       str(g(row, "OS according to the configuration file") or "").strip(),
            "datacenter":      str(g(row, "Datacenter") or "").strip(),
            "vcenter_sdk":     str(g(row, "VI SDK Server") or "").strip(),
            "annotation":      str(g(row, "Annotation") or "").strip(),
            "power_on":        _str_date(g(row, "PowerOn")),
            "creation_date":   _str_date(g(row, "Creation date")),
            "hw_version":      _int(g(row, "HW version")),
            "firmware":        str(g(row, "Firmware") or "").strip(),
            "connection_state": str(g(row, "Connection state") or "").strip(),
            "guest_state":     str(g(row, "Guest state") or "").strip(),
            "cbt":             str(g(row, "CBT") or "").strip(),
            "vm_id":           str(g(row, "VM ID") or "").strip(),
            "path":            str(g(row, "Path") or "").strip(),
        }
        vms.append(vm)
    return vms


def _parse_vhost_sheet(ws) -> list:
    """Parse vHost worksheet → list of ESXi host dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    def g(row, col_name):
        i = col_idx.get(col_name)
        return row[i] if i is not None and i < len(row) else None

    hosts = []
    for row in rows[1:]:
        name = g(row, "Host")
        if not name:
            continue
        hosts.append({
            "name":          str(name),
            "datacenter":    str(g(row, "Datacenter") or ""),
            "cluster":       str(g(row, "Cluster") or ""),
            "cpu_model":     str(g(row, "CPU Model") or ""),
            "cpu_sockets":   _int(g(row, "# CPU")),
            "cpu_cores":     _int(g(row, "# Cores")),
            "cpu_threads":   _int(g(row, "# Threads")),
            "ram_mb":        _int(g(row, "# Memory")),
            "vms":           _int(g(row, "# VMs")),
            "vcenter_sdk":   str(g(row, "VI SDK Server") or ""),
        })
    return hosts


def _parse_vdisk_sheet(ws) -> list:
    """Parse vDisk worksheet → list of disk dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    def g(row, col_name):
        i = col_idx.get(col_name)
        return row[i] if i is not None and i < len(row) else None

    disks = []
    for row in rows[1:]:
        vm_name = g(row, "VM")
        if not vm_name:
            continue
        disks.append({
            "vm":            str(vm_name),
            "disk_label":    str(g(row, "Disk") or ""),
            "capacity_mb":   _int(g(row, "Capacity MB")),
            "format":        str(g(row, "Disk Format") or ""),
            "datastore":     str(g(row, "Datastore") or ""),
            "vcenter_sdk":   str(g(row, "VI SDK Server") or ""),
        })
    return disks


def _parse_vdatastore_sheet(ws) -> list:
    """Parse vDatastore worksheet."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    def g(row, col_name):
        i = col_idx.get(col_name)
        return row[i] if i is not None and i < len(row) else None

    ds_list = []
    for row in rows[1:]:
        name = g(row, "Name")
        if not name:
            continue
        ds_list.append({
            "name":          str(name),
            "type":          str(g(row, "Type") or ""),
            "capacity_mb":   _int(g(row, "Capacity MB")),
            "freespace_mb":  _int(g(row, "Free Space MB")),
            "datacenter":    str(g(row, "Datacenter") or ""),
            "cluster":       str(g(row, "Cluster") or ""),
            "vcenter_sdk":   str(g(row, "VI SDK Server") or ""),
        })
    return ds_list


def _int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def _str_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v).strip() or None


def parse_rvtools_xlsx(filepath: str) -> dict:
    """
    Parse an RVTools XLSX file.
    Returns dict with keys: vms, hosts, disks, datastores, source_file, parsed_at
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {"error": str(e), "source_file": filepath}

    result = {
        "source_file":  filepath,
        "parsed_at":    datetime.datetime.now().isoformat(),
        "vms":          [],
        "hosts":        [],
        "disks":        [],
        "datastores":   [],
    }

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sl = sheet.lower()
        if sl == "vinfo":
            result["vms"] = _parse_vinfo_sheet(ws)
        elif sl == "vhost":
            result["hosts"] = _parse_vhost_sheet(ws)
        elif sl in ("vdisk",):
            result["disks"] = _parse_vdisk_sheet(ws)
        elif sl in ("vdatastore",):
            result["datastores"] = _parse_vdatastore_sheet(ws)

    wb.close()
    return result


# ── Summary builder ───────────────────────────────────────────────────────────
def build_summary(parsed: dict) -> dict:
    """Derive aggregate summary from parsed RVTools data."""
    vms = parsed.get("vms", [])

    # Power states
    power_counts = {}
    for vm in vms:
        ps = vm.get("powerstate", "Unknown") or "Unknown"
        power_counts[ps] = power_counts.get(ps, 0) + 1

    # OS distribution
    os_counts = {}
    for vm in vms:
        os_str = vm.get("os_tools") or vm.get("os_config") or "Unknown"
        # Bucket into broad categories
        ol = os_str.lower()
        if "windows" in ol:
            key = _windows_version(os_str)
        elif "rhel" in ol or "red hat" in ol:
            key = "RHEL"
        elif "centos" in ol:
            key = "CentOS"
        elif "ubuntu" in ol:
            key = "Ubuntu"
        elif "debian" in ol:
            key = "Debian"
        elif "sles" in ol or "suse" in ol:
            key = "SUSE / SLES"
        elif "oracle" in ol:
            key = "Oracle Linux"
        elif "esxi" in ol or "vmware" in ol:
            key = "VMware ESXi"
        elif not os_str or os_str == "None":
            key = "Unknown"
        else:
            key = os_str[:40]
        os_counts[key] = os_counts.get(key, 0) + 1

    # Resource totals (powered-on only)
    powered_on = [v for v in vms if (v.get("powerstate") or "").lower() == "poweredon"]
    total_cpu = sum(v.get("cpus", 0) or 0 for v in powered_on)
    total_ram_gb = round(sum(v.get("memory_mb", 0) or 0 for v in powered_on) / 1024, 1)
    # Use disk_capacity_mb (Total disk capacity MiB) when available, fall back to provisioned_mb
    def _prov(v):
        d = v.get("disk_capacity_mb", 0) or 0
        p = v.get("provisioned_mb", 0) or 0
        return d if d > 0 else p
    total_prov_tb = round(sum(_prov(v) for v in vms) / (1024 * 1024), 2)
    total_used_tb = round(sum(v.get("inuse_mb", 0) or 0 for v in vms) / (1024 * 1024), 2)

    # Clusters
    clusters = sorted(set(v.get("cluster", "") for v in vms if v.get("cluster")))
    datacenters = sorted(set(v.get("datacenter", "") for v in vms if v.get("datacenter")))

    # Top resource consumers (powered-on, by RAM)
    top10_ram = sorted(
        [v for v in powered_on if v.get("memory_mb")],
        key=lambda v: v.get("memory_mb", 0),
        reverse=True
    )[:10]

    hosts = parsed.get("hosts", [])
    datastores = parsed.get("datastores", [])

    ds_total_cap_tb = round(sum(d.get("capacity_mb", 0) or 0 for d in datastores) / (1024 * 1024), 2)
    ds_total_free_tb = round(sum(d.get("freespace_mb", 0) or 0 for d in datastores) / (1024 * 1024), 2)

    return {
        "total_vms":          len(vms),
        "templates":          sum(1 for v in vms if v.get("template")),
        "powered_on":         power_counts.get("poweredOn", 0),
        "powered_off":        power_counts.get("poweredOff", 0),
        "suspended":          power_counts.get("suspended", 0),
        "power_counts":       power_counts,
        "os_counts":          dict(sorted(os_counts.items(), key=lambda x: x[1], reverse=True)),
        "total_vcpu":         total_cpu,
        "total_ram_gb":       total_ram_gb,
        "total_provisioned_tb": total_prov_tb,
        "total_used_tb":      total_used_tb,
        "clusters":           clusters,
        "datacenters":        datacenters,
        "total_hosts":        len(hosts),
        "total_datastores":   len(datastores),
        "ds_capacity_tb":     ds_total_cap_tb,
        "ds_free_tb":         ds_total_free_tb,
        "top10_by_ram":       [
            {"name": v["name"], "cpus": v.get("cpus", 0), "ram_gb": round((v.get("memory_mb", 0) or 0) / 1024, 1)}
            for v in top10_ram
        ],
    }


def _windows_version(os_str: str) -> str:
    ol = os_str.lower()
    if "2022" in ol:
        return "Windows Server 2022"
    if "2019" in ol:
        return "Windows Server 2019"
    if "2016" in ol:
        return "Windows Server 2016"
    if "2012" in ol:
        return "Windows Server 2012"
    if "2008" in ol:
        return "Windows Server 2008"
    if "10" in ol:
        return "Windows 10/11"
    if "7" in ol:
        return "Windows 7"
    return "Windows (other)"


# ── Report scanner ────────────────────────────────────────────────────────────
def _file_mtime_iso(path: str) -> str:
    try:
        return datetime.datetime.fromtimestamp(os.path.getmtime(path)).isoformat()
    except Exception:
        return None


def scan_reports(extra_dirs: list = None) -> list:
    """
    Scan all known directories for RVTools XLSX files.
    Returns list of report metadata dicts (not full vm lists).
    """
    dirs = list(SCAN_DIRS)
    if extra_dirs:
        dirs.extend(extra_dirs)

    seen_canonical = set()
    seen_basename = {}   # basename -> already-added path (prefer EXPORTS_DIR)
    reports = []
    for d in dirs:
        if not os.path.isdir(d):
            continue
        for f in glob.glob(os.path.join(d, "*.xlsx")):
            canonical = os.path.normcase(os.path.abspath(f))
            if canonical in seen_canonical:
                continue
            seen_canonical.add(canonical)
            bname = os.path.basename(f).lower()
            if bname in seen_basename:
                # Keep the one from EXPORTS_DIR, skip the other
                existing = seen_basename[bname]
                if os.path.normcase(d) == os.path.normcase(EXPORTS_DIR):
                    # Remove old entry and replace with this one
                    reports = [r for r in reports if os.path.basename(r["file"]).lower() != bname]
                else:
                    continue
            seen_basename[bname] = f
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                sheets = wb.sheetnames
                wb.close()
                if "vInfo" not in sheets:
                    continue   # Not an RVTools file
                reports.append({
                    "file":       f,
                    "filename":   os.path.basename(f),
                    "dir":        d,
                    "modified":   _file_mtime_iso(f),
                    "size_kb":    round(os.path.getsize(f) / 1024, 1),
                    "sheets":     sheets,
                })
            except Exception:
                pass
    return sorted(reports, key=lambda r: r.get("modified") or "", reverse=True)


# ── High-level API helpers ─────────────────────────────────────────────────────
def get_all_reports_with_summary() -> list:
    """
    Scan, parse and summarise ALL found RVTools reports.
    Returns list of {file, modified, vcenter_guess, summary, ...}
    """
    reports = scan_reports()
    result = []
    for r in reports:
        parsed  = parse_rvtools_xlsx(r["file"])
        summary = build_summary(parsed)
        # Guess which vCenter this file belongs to
        vcenter_guess = _guess_vcenter(r["file"], parsed.get("vms", []))
        result.append({
            **r,
            "vcenter_host": vcenter_guess.get("host"),
            "vcenter_name": vcenter_guess.get("name"),
            "summary":      summary,
        })
    return result


def get_report_vms(filepath: str) -> dict:
    """Full VM list from a specific report file."""
    parsed  = parse_rvtools_xlsx(filepath)
    summary = build_summary(parsed)
    return {
        "source_file": filepath,
        "parsed_at":   parsed.get("parsed_at"),
        "summary":     summary,
        "vms":         parsed.get("vms", []),
        "hosts":       parsed.get("hosts", []),
        "datastores":  parsed.get("datastores", []),
    }


def run_and_get_report(vcenter_id: str) -> dict:
    """
    Run RVTools for a specific vCenter (by host IP or name),
    then parse and return the summary.
    """
    vc = next((v for v in VCENTERS if v["host"] == vcenter_id or v["name"] == vcenter_id), None)
    if not vc:
        return {"success": False, "message": f"vCenter '{vcenter_id}' not found in configuration."}

    run_result = run_rvtools_for_vcenter(vc)
    if not run_result["success"]:
        return run_result

    parsed  = parse_rvtools_xlsx(run_result["file"])
    summary = build_summary(parsed)
    return {
        "success":     True,
        "message":     run_result["message"],
        "file":        run_result["file"],
        "vcenter_host": vc["host"],
        "vcenter_name": vc["name"],
        "summary":     summary,
    }


def _guess_vcenter(filepath: str, vms: list) -> dict:
    """
    Try to match a report to a known vCenter.
    Priority: 1) vCenter IP in filename  2) VI SDK Server from VM rows  3) empty
    """
    fname = os.path.basename(filepath)
    for vc in VCENTERS:
        if vc["host"].replace(".", r"\.") and re.search(re.escape(vc["host"]), fname):
            return vc
        if vc["name"] and vc["name"].lower() in fname.lower():
            return vc

    # Check VI SDK Server from first VM
    if vms:
        sdk = vms[0].get("vcenter_sdk", "")
        for vc in VCENTERS:
            if vc["host"] == sdk:
                return vc
        if sdk:
            return {"host": sdk, "name": sdk}

    return {"host": None, "name": "Unknown"}


def get_rvtools_status() -> dict:
    """Return installation status and quick stats."""
    exe = find_rvtools_exe()
    reports = scan_reports()
    return {
        "installed":       exe is not None,
        "exe_path":        exe,
        "exports_dir":     EXPORTS_DIR,
        "reports_found":   len(reports),
        "reports":         [{"file": r["file"], "modified": r["modified"], "size_kb": r["size_kb"]} for r in reports],
        "vcenters":        [{"host": v["host"], "name": v["name"]} for v in VCENTERS],
    }
