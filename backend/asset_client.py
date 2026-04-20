"""
Asset Management Client — parse rack inventory Excel, ping IPs, execute OEM power actions.
Supports HP iLO, Dell iDRAC, Cisco CIMC, Quanta/SuperMicro via Redfish.
"""
import os, socket, subprocess, json, copy
from pathlib import Path
from typing import Any, Optional
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import requests
import openpyxl

DATA_DIR = Path(__file__).parent / "data"
DC_FILE  = DATA_DIR / "dc_inventory.xlsx"
DR_FILE  = DATA_DIR / "dr_inventory.xlsx"

# ── Column header normalisation ────────────────────────────────────────────────
_HDR_MAP = {
    "rack number"        : "rack_number",
    "e"                  : "rack_number",   # Rack-10 typo
    "assets position"    : "position",
    "physical assets name": "asset_name",
    "serial number"      : "serial",
    "serial number "     : "serial",
    "serial no"          : "serial",
    "model"              : "model",
    "mgmt port"          : "mgmt_port",
    "mgmt ip"            : "mgmt_ip",
    "data ports"         : "data_ports",
    "os / hypervisor"    : "os_hypervisor",
    "os name"            : "os_hypervisor",
    "os name/version"    : "os_hypervisor",
    "os"                 : "os_hypervisor",
    "os ip"              : "os_ip",
    "remarks"            : "remarks",
    "power state"        : "power_state",
    "owner"              : "owner",
    "hostname"           : "hostname",
    "switch name/model no": "switch_info",
}

def _norm_hdr(h: Any) -> str:
    if h is None:
        return ""
    return _HDR_MAP.get(str(h).strip().lower(), str(h).strip().lower().replace(" ", "_"))


def _clean(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # Remove non-breaking spaces
    return s.replace("\xa0", " ").replace("\n", " | ")


def parse_inventory(filepath: Path) -> dict:
    """Parse an Excel inventory file.
    Returns { "racks": { "Rack-01": [asset, ...], ... }, "sheet_order": [...] }
    """
    if not filepath.exists():
        return {"racks": {}, "sheet_order": []}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    racks: dict = {}
    sheet_order: list = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Skip empty sheets
        non_empty = [r for r in rows if any(v for v in r if v is not None)]
        if len(non_empty) < 2:
            continue

        # First non-empty row = headers
        header_row = non_empty[0]
        headers = [_norm_hdr(h) for h in header_row]

        # For sheets like SR-05 where the OS-name column has no header but sits
        # immediately after os_ip, map it to os_hypervisor automatically.
        if "os_hypervisor" not in headers:
            for i, h in enumerate(headers):
                if h == "" and i > 0 and headers[i - 1] == "os_ip":
                    headers[i] = "os_hypervisor"
                    break

        assets: list = []
        for raw_row in non_empty[1:]:
            # Skip rows where all meaningful cells are empty
            if not any(v for v in raw_row if v is not None):
                continue
            asset: dict = {}
            for idx, key in enumerate(headers):
                if not key:
                    continue
                val = raw_row[idx] if idx < len(raw_row) else None
                cleaned = _clean(val)
                if key in ("owner", "hostname") and key in asset:
                    # Merge owner / hostname into one field
                    if cleaned:
                        asset["owner"] = (asset.get("owner", "") + " " + cleaned).strip()
                else:
                    asset[key] = cleaned

            # Ensure all standard keys exist
            for k in ("rack_number", "position", "asset_name", "serial", "model",
                      "mgmt_port", "mgmt_ip", "data_ports", "os_hypervisor",
                      "os_ip", "remarks", "power_state", "owner"):
                asset.setdefault(k, "")

            # Skip section-header rows (e.g. the "SWITCHES" divider row
            # that has a rack_number value but no asset name)
            if not asset.get("asset_name", "").strip():
                continue

            # Generate a stable row ID
            asset["_id"] = f"{sheet_name}___{asset.get('position','')}__{asset.get('serial','')}"
            assets.append(asset)

        if assets:
            racks[sheet_name] = assets
            sheet_order.append(sheet_name)

    return {"racks": racks, "sheet_order": sheet_order}


def save_inventory(filepath: Path, data: dict) -> bool:
    """Write modified inventory back to Excel (preserving style as plain xlsx)."""
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        standard_headers = [
            ("Rack Number",       "rack_number"),
            ("Assets Position",   "position"),
            ("Physical Assets Name", "asset_name"),
            ("Serial Number",     "serial"),
            ("Model",             "model"),
            ("Mgmt Port",         "mgmt_port"),
            ("Mgmt IP",           "mgmt_ip"),
            ("Data Ports",        "data_ports"),
            ("OS / Hypervisor",   "os_hypervisor"),
            ("OS IP",             "os_ip"),
            ("Remarks",           "remarks"),
            ("POWER STATE",       "power_state"),
            ("OWNER",             "owner"),
        ]

        for sheet_name in data.get("sheet_order", []):
            assets = data["racks"].get(sheet_name, [])
            ws = wb.create_sheet(sheet_name)
            ws.append([h[0] for h in standard_headers])
            for asset in assets:
                ws.append([asset.get(h[1], "") for h in standard_headers])

        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"[AssetClient] save_inventory error: {e}")
        return False


# ── IP Reachability ────────────────────────────────────────────────────────────
def ping_ip(ip: str, timeout: float = 1.5) -> bool:
    """Check if an IP is reachable. Uses HTTPS port first (fast), then ICMP."""
    if not ip:
        return False
    ip = ip.strip().split("/")[0].split("[")[0].strip()
    if not ip or ip.lower() in ("none", "unknown", "na", "", "no mgmt"):
        return False
    # Validate IP-ish
    parts = ip.split(".")
    if len(parts) != 4:
        return False
    try:
        for port in (443, 80, 22, 623):
            try:
                s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                s.settimeout(timeout)
                r = s.connect_ex((ip, port))
                s.close()
                if r == 0:
                    return True
            except Exception:
                pass
        # ICMP ping fallback (Windows)
        result = subprocess.run(
            ["ping", "-n", "1", "-w", "1500", ip],
            capture_output=True, timeout=4
        )
        return result.returncode == 0
    except Exception:
        return False


def ping_many(ips: list[str]) -> dict[str, bool]:
    """Check reachability of multiple IPs concurrently."""
    import concurrent.futures
    unique = list({ip for ip in ips if ip and ip.strip()})
    with concurrent.futures.ThreadPoolExecutor(max_workers=30) as ex:
        results = dict(zip(unique, ex.map(ping_ip, unique)))
    return results


# ── OEM Detection ──────────────────────────────────────────────────────────────
def detect_oem(asset_name: str, model: str) -> str:
    text = f"{asset_name or ''} {model or ''}".lower()
    if any(x in text for x in ["hp proliant", "hp synergy", "hpe ", "hp enclo", "hp 3par",
                                 "nimble", "alletra", "proliant", "storeserv", "synergy"]):
        return "hp"
    if any(x in text for x in ["dell", "poweredge", "apex cloud"]):
        return "dell"
    if any(x in text for x in ["cisco", "hyperflex", "ucs"]):
        return "cisco"
    if "supermicro" in text:
        return "supermicro"
    if "quanta" in text:
        return "quanta"
    return "generic"


# ── Redfish Power Action ───────────────────────────────────────────────────────
_RESET_MAP = {
    "poweron"    : "On",
    "reboot"     : "GracefulRestart",
    "shutdown"   : "GracefulShutdown",
    "forceoff"   : "ForceOff",
    "forcereboot": "ForceRestart",
}

# Known Redfish system paths by OEM
_OEM_PATHS = {
    "hp"        : "/redfish/v1/Systems/1",
    "dell"      : "/redfish/v1/Systems/System.Embedded.1",
    "supermicro": "/redfish/v1/Systems/1",
    "quanta"    : "/redfish/v1/Systems/1",
}


def _get_session(username: str, password: str) -> requests.Session:
    sess = requests.Session()
    sess.verify = False
    sess.auth   = (username, password)
    sess.headers.update({"Content-Type": "application/json", "Accept": "application/json"})
    return sess


def _discover_system_url(ip: str, sess: requests.Session) -> Optional[str]:
    """Auto-discover the first system member URL via Redfish."""
    try:
        r = sess.get(f"https://{ip}/redfish/v1/Systems/", timeout=6)
        data = r.json()
        members = data.get("Members", [])
        if members:
            return members[0].get("@odata.id", "")
    except Exception:
        pass
    return None


def power_action(mgmt_ip: str, action: str, username: str, password: str,
                 asset_name: str = "", model: str = "") -> dict:
    """
    Execute a power action on a physical server via Redfish.
    Returns {"success": bool, "message": str}
    """
    ip = (mgmt_ip or "").strip().split("/")[0]
    if not ip:
        return {"success": False, "message": "No management IP available"}

    reset_type = _RESET_MAP.get(action.lower(), "GracefulRestart")
    oem        = detect_oem(asset_name, model)
    sess       = _get_session(username, password)

    # Get system base URL
    system_path = _OEM_PATHS.get(oem)
    if not system_path:
        system_path = _discover_system_url(ip, sess)
    if not system_path:
        return {"success": False, "message": f"Cannot locate Redfish system for OEM '{oem}' at {ip}"}

    action_url = f"https://{ip}{system_path}/Actions/ComputerSystem.Reset"

    try:
        r = sess.post(action_url, json={"ResetType": reset_type}, timeout=12)
        if r.status_code in (200, 202, 204):
            return {"success": True,
                    "message": f"'{action}' command accepted by {ip} (HTTP {r.status_code})"}
        # Try alternate path for Cisco
        if oem == "cisco" and r.status_code in (400, 404, 405):
            sys_url = _discover_system_url(ip, sess)
            if sys_url:
                alt_url = f"https://{ip}{sys_url}/Actions/ComputerSystem.Reset"
                r2 = sess.post(alt_url, json={"ResetType": reset_type}, timeout=12)
                if r2.status_code in (200, 202, 204):
                    return {"success": True, "message": f"'{action}' accepted (Cisco, HTTP {r2.status_code})"}
        return {"success": False,
                "message": f"HTTP {r.status_code}: {r.text[:300] if r.text else 'No body'}"}
    except requests.exceptions.ConnectTimeout:
        return {"success": False, "message": f"Connection timeout to {ip} — management interface unreachable"}
    except requests.exceptions.ConnectionError as e:
        return {"success": False, "message": f"Connection error: {str(e)[:200]}"}
    except Exception as e:
        return {"success": False, "message": f"Error: {str(e)[:300]}"}
